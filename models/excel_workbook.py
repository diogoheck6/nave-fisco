from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
import pandas as pd
from datetime import datetime

def is_file_open(filename):
    """Verifica se um arquivo está aberto tentando renomeá-lo temporariamente."""
    if not os.path.exists(filename):
        return False

    try:
        os.rename(filename, filename)  # Renomeia o arquivo consigo mesmo
        return False
    except OSError:
        return True
    

class ExcelWorkbook:
    def __init__(self, competencia, filename=None):

        """Inicializa o nome do arquivo, mas não cria ou carrega o Workbook imediatamente."""
        if not filename:
            filename = "workbook_default.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'

        self.filename = filename
        self.workbook = None
        self.competencia = competencia

        # Define estilos numéricos e de data
        self.number_style_defined = False
        self.date_style_defined = False
        self._define_styles()

    def _define_styles(self):
        """Definir uma vez os estilos que serão aplicados às células."""
        self.number_style = NamedStyle(
            name="number_style",
            # number_format='#,##0.00;[Red]-#,##0.00',
            number_format='_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
            alignment=Alignment(horizontal='right')
        )

        self.date_style = NamedStyle(
            name="date_style",
            number_format='DD/MM/YYYY',
            alignment=Alignment(horizontal='right')
        )

    def create_new_workbook(self):
        """Cria um novo Workbook."""
        self.workbook = Workbook()
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        self._add_styles_to_workbook()
        # print("Novo workbook criado e planilha padrão removida.")

    def _add_styles_to_workbook(self):
        """Adiciona os NamedStyles ao Workbook, garantindo que não sejam duplicados."""
        if not self.number_style_defined:
            self.workbook.add_named_style(self.number_style)
            self.number_style_defined = True

        if not self.date_style_defined:
            self.workbook.add_named_style(self.date_style)
            self.date_style_defined = True

    def load_existing_workbook(self):
        """Carrega um Workbook existente pelo nome do arquivo."""
        if os.path.exists(self.filename):
            self.workbook = load_workbook(self.filename)
            # print(f"Workbook '{self.filename}' carregado.")
            self._add_styles_to_workbook()
        else:
            raise FileNotFoundError(f"O arquivo '{self.filename}' não existe.")

    def add_sheet(self, sheet_name):
        """Adiciona uma nova planilha ao Workbook e define opções padrões."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        if sheet_name in self.workbook.sheetnames:
            # print(f"A planilha '{sheet_name}' já existe.")
            pass
        else:
            self.workbook.create_sheet(title=sheet_name)
            self.set_tab_color(sheet_name, "green")
            self.toggle_gridlines(sheet_name, show_gridlines=False)
            # print(f"Planilha '{sheet_name}' adicionada.")

    def remove_sheet(self, sheet_name):
        """Remove uma planilha do Workbook."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.workbook.remove(sheet)
            # print(f"Planilha '{sheet_name}' removida.")
        else:
            # print(f"A planilha '{sheet_name}' não existe.")
            pass

    def set_tab_color(self, sheet_name, color):
        """Define a cor da aba de uma planilha."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        
        color_dict = {
            'red': 'FF0000',
            'yellow': 'FFFF00',
            'green': '00FF00',
        }
        
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if color in color_dict:
                sheet.sheet_properties.tabColor = color_dict[color]
                # print(f"Cor da aba da planilha '{sheet_name}' definida como {color}.")
            else:
                # print(f"Cor '{color}' não está definida. Use 'red', 'yellow', ou 'green'.")
                pass
        else:
            # print(f"A planilha '{sheet_name}' não existe.")
            pass

    def fechamento_sem_diferenca(self, sheet, tipo):
        """
        Adiciona uma mensagem à planilha indicando que todas as notas do SAT também constam no Questor,
        em caso de ausência de diferenças.
        """
        # Definindo a mensagem com base no tipo fornecido
        if tipo.lower() == 'entrada':
            mensagem = "TODAS AS NOTAS DE ENTRADAS QUE CONSTAM NO SAT TAMBÉM CONSTAM NO QUESTOR"
        else:
            mensagem = "TODAS AS NOTAS DE SAÍDAS QUE CONSTAM NO SAT TAMBÉM CONSTAM NO QUESTOR"

        # Determinando a próxima linha vazia
        max_row = sheet.max_row + 2  # Adiciona uma folga de 2 linhas após os dados existentes

        # Inserindo a mensagem
        cell = sheet.cell(row=max_row, column=1, value=mensagem)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center')
    
    def inserir_mensagem_sem_diferenca(self, sheet_name, tipo):
        """
        Insere uma mensagem em negrito com tamanho de fonte 16 em uma planilha existente,
        indicando que todas as notas do SAT também constam no Questor.

        :param sheet_name: O nome da planilha onde a mensagem será inserida.
        :param tipo: Define se a mensagem é para 'entrada' ou 'saida'.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe no workbook.")

        sheet = self.workbook[sheet_name]

        # Define a mensagem com base no tipo especificado
        if tipo.lower() == 'entrada':
            mensagem = "TODAS AS NOTAS DE ENTRADAS QUE CONSTAM NO SAT TAMBÉM CONSTAM NO QUESTOR"
        elif tipo.lower() == 'questor':
            mensagem = "TODAS AS NOTAS DE ENTRADAS QUE CONSTAM NO QUESTOR TAMBÉM CONSTAM NO SAT"
        elif tipo.lower() == 'saidas questor':
            mensagem = "TODAS AS NOTAS DE SAÍDAS QUE CONSTAM NO QUESTOR TAMBÉM CONSTAM NO SAT"
        elif tipo.lower() == 'msg_dif_saidas':
            mensagem = 'NÃO HÁ DIFERENÇAS ENTRE VALOR CONTABIL, BASE DE CÁLCULO ICMS, VALOR ICMS E VALOR IPI ENTRE AS NOTAS DE SAÍDA'
        elif tipo.lower() == 'msg_dif_entradas':
            mensagem = 'NÃO HÁ DIFERENÇAS ENTRE VALOR CONTABIL, BASE DE CÁLCULO ICMS, VALOR ICMS E VALOR IPI ENTRE AS NOTAS DE ENTRADA'
        else:
            mensagem = "TODAS AS NOTAS DE SAÍDAS QUE CONSTAM NO SAT TAMBÉM CONSTAM NO QUESTOR"

        # Inserindo a mensagem na célula A1
        cell = sheet.cell(row=5, column=2, value=mensagem)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='left')
        
        # Ajusta a largura da coluna para acomodar a mensagem
        # sheet.column_dimensions['A'].width = 70

    def populate_sheet(self, sheet_name, data, header=True, footer=None):
        """Popula uma planilha com dados, incluindo opcionalmente um cabeçalho e rodapé."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")

        sheet = self.workbook[sheet_name]

        if header and len(data) > 0:
            # Inserir cabeçalho
            cols = list(data[0].keys())
            for col_num, column_title in enumerate(cols, start=1):
                cell = sheet.cell(row=1, column=col_num, value=column_title.upper())
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        # Popula os dados
        for row_num, entry in enumerate(data, start=2):
            for col_num, (key, value) in enumerate(entry.items(), start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Inserir rodapé, se houver
        if footer:
            max_row = len(data) + 2
            for col_num, (key, value) in enumerate(footer.items(), start=1):
                cell = sheet.cell(row=max_row, column=col_num, value=value)
                cell.font = Font(italic=True)

        # print(f"Planilha '{sheet_name}' populada com dados.")

    def centralize_non_numeric_texts(self, sheet_name, start_row=2):
        """Centraliza os textos que não são numéricos."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]

        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                if not isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='center')
    
    def format_numbers(self, sheet_name, start_row=2):
        """Formata números e datas nas células da planilha."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]

        # Lista de nomes de colunas que devem ser tratadas como numéricas
        colunas_numericas = ['VLR NF', 'VLR BC ICMS', 'VLR ICMS', 'VLR TOTAL', 
                             'BC ICMS', 'ICMS', 'VLR IPI', 'IPI', 'VLR NF_SAT',
                             'VLR NF_QUESTOR', 'DIF VLR NF', 'VLR BC ICMS_SAT',
                             'VLR BC ICMS_QUESTOR', 'DIF VLR BC ICMS',
                             'VLR ICMS_SAT', 'VLR ICMS_QUESTOR', 'DIF VLR ICMS',
                             'VLR IPI_SAT', 'VLR IPI_QUESTOR', 'DIF VLR IPI']

        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
            for cell in row:
                # Obtém o nome da coluna para a célula atual
                coluna_nome = sheet.cell(row=1, column=cell.column).value or ''

                # Verifica se a célula pertence a uma coluna numérica antes de processar
                if coluna_nome in colunas_numericas:
                    try:
                        # Remove as vírgulas e converte para float
                        value_str = str(cell.value).replace(',', '')
                        # cell_value = float(value_str) if value_str.replace('.', '', 1).isdigit() else cell.value
                        cell_value = float(value_str) 
                        
                        # Formata o valor para visualização
                        if isinstance(cell_value, (int, float)):
                            # if cell_value == 0:
                            #     cell.value = 0.0
                            #     cell.style = self.number_style
                            #     cell.alignment = Alignment(horizontal='right')
                            # else:
                            cell.value = cell_value
                            cell.style = self.number_style
                    except ValueError:
                        # Se não for possível converter, mantenha o valor original
                        pass

                # Formatando datas nas colunas identificadas pelo nome
                if 'DATA' in coluna_nome.upper():
                    try:
                        if isinstance(cell.value, str):
                            date_value = pd.to_datetime(cell.value, dayfirst=True, errors='coerce')
                            if not pd.isnull(date_value):
                                cell.value = date_value.strftime('%d/%m/%Y')
                                cell.style = self.date_style
                        elif isinstance(cell.value, datetime):
                            cell.value = cell.value.strftime('%d/%m/%Y')
                            cell.style = self.date_style
                    except Exception:
                        pass

    def apply_header_filter(self, sheet_name):
        """Aplica filtros no cabeçalho da planilha."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]
        max_column = sheet.max_column
        last_column_letter = get_column_letter(max_column)
        
        sheet.auto_filter.ref = f"A1:{last_column_letter}1"
        # print(f"Filtros aplicados no cabeçalho da planilha '{sheet_name}'.")

    def apply_zebra_stripes(self, sheet_name):
        """Aplica estilos de zebra cinza claro nas linhas pares da planilha."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = gray_fill

        # print(f"Estilos de zebra aplicados na planilha '{sheet_name}'.")

    def auto_adjust_columns(self, sheet_name):
        """Ajusta a largura das colunas automaticamente com base no conteúdo."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 4)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # print(f"Largura das colunas ajustada automaticamente para a planilha '{sheet_name}'.")

    def populate_closing_report(self, sheet_name, fechamento):
        """
        Popula uma planilha com base em um fechamento (como no formato especificado).
        As células recebem o conteúdo formatado e seguem o espaçamento correto.
        """
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado.")
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Planilha '{sheet_name}' não existe.")
        
        sheet = self.workbook[sheet_name]

        row_num = 1  # Começa na primeira linha
        for entry in fechamento:
            if entry:  # Preenche os dados formatados
                for col, value in entry.items():
                    col_index = ["A", "B", "C", "D"].index(col) + 1
                    cell = sheet.cell(row=row_num, column=col_index, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if col in ["B", "C", "D"] and isinstance(value, str) and value.replace(",", "").replace(".", "").isdigit():
                        cell.number_format = '#,##0.00'
            row_num += 1  # Próxima linha
        # print(f"Planilha '{sheet_name}' populada com o fechamento.")

    
    

    def save(self):
        """Salva o Workbook no arquivo especificado."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        
        if not hasattr(self, 'competencia') or not self.competencia:
            raise ValueError("A 'competencia' não foi especificada.")
        
        # Verificar se o diretório base do arquivo existe
        base_directory = os.path.dirname(self.filename)
        

        
        if not os.path.exists(base_directory):
            raise FileNotFoundError(f"O diretório base não existe: {base_directory}")
        
        # Criar um subdiretório para a competência
        competencia_directory = os.path.join(base_directory, self.competencia)
        
        # Criar o diretório da competência se ele não existir
        if not os.path.exists(competencia_directory):
            try:
                os.makedirs(competencia_directory)
                print(f"Diretório criado: {competencia_directory}")
            except OSError as e:
                raise OSError(f"Erro ao criar o diretório da competência: {e}")
        
        # Tente salvar o arquivo com o nome base fornecido ou nomes incrementais, se necessário
        base_filename, file_extension = os.path.splitext(os.path.basename(self.filename))
        attempt = 0
        
        while True:
            try:
                # Use o nome original fornecido para construir o caminho completo
                save_filename = os.path.join(competencia_directory, f"{base_filename}{file_extension}")
                
                # Acrescente sufixos ao nome do arquivo caso ele esteja aberto
                if attempt > 0:
                    save_filename = os.path.join(competencia_directory, f"{base_filename}_v{attempt}{file_extension}")
                
                if not is_file_open(save_filename):
                    # Salvar o arquivo no nome e caminho especificados ou com sufixo
                    self.workbook.save(save_filename)
                    print(f"Arquivo salvo em: {save_filename}")
                    break

            except Exception as e:
                print(f"Erro ao tentar salvar o arquivo: {e}")
            
            # Incrementar o sufixo para tentar com outro nome
            attempt += 1
        


    def toggle_gridlines(self, sheet_name, show_gridlines=False):
        """Ativa ou desativa as linhas de grade de uma planilha."""
        if not self.workbook:
            raise ValueError("Nenhum workbook foi carregado ou criado ainda.")
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet.sheet_view.showGridLines = show_gridlines
            # print(f"Linhas de grade {'ativadas' if show_gridlines else 'desativadas'} na planilha '{sheet_name}'.")
        else:
            # print(f"A planilha '{sheet_name}' não existe.")
            pass


if __name__ == '__main__':

    pass